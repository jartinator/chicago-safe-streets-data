"""Historical CDOT bikeway mileage, recovered from FOIA S145367-071326.

The public CDOT Bike Routes layer (`hvv9-38ut`) is current-state only, and the
quarterly Bike Lane Mileage Tracker on the Complete Streets page is overwritten
in place — so OYL could only ever build mileage history *forward*, from the
snapshots under `data/snapshots/`, which start 2026-07-11.

CDOT's response to FOIA S145367-071326 (released 2026-07-24) supplies the history
directly. The output of this module is spliced onto the front of the published
series by `aggregate.build_bikeway_mileage_series`, and drives the delivered-miles
ledger in `commitments_metrics` (DECISIONS.md #35-#36). Two records in it matter:

1. **`CompleteStreets_Dashboard.xlsx`, sheet `R_Dashboard`.** CDOT's own program
   dashboard, keyed by stable `R_*` row labels, holding annual centerline miles by
   facility type for **2010 through 2025** — both the standing network at each
   year end and the miles installed during that year. This is the authoritative
   series and the one this module publishes.

2. **`GIS/Bikeway_Network_2024_Final.shp`.** The internal bikeway layer, which
   *does* carry `BW_INST_YR` / `BW_INST_MO` per segment — the field the public
   layer lacks. Grouped by install year it gives a segment-level view of when the
   surviving 2024 network was built.

Read the caveat on `segment_install_years` before treating (2) as a second
history: it is a snapshot of survivors, not a record of the past.

Run directly to refresh the committed output:

    python pipeline/foia_bikeway_history.py
"""
import json
from collections import defaultdict

from config import (
    CDOT_BIKEWAY_2024_LAYER,
    CDOT_BIKEWAY_HISTORY_PATH,
    CDOT_COMPLETE_STREETS_DASHBOARD,
    FACILITY_CATEGORIES,
)

# Sheet holding the machine-readable form of the dashboard. The sibling
# `Dashboard` sheet is the same numbers laid out for humans (merged title rows,
# mayoral-term banners); `R_Dashboard` keys every row with a stable `R_*` label,
# so it survives CDOT re-laying-out the presentation sheet.
DASHBOARD_SHEET = "R_Dashboard"
YEAR_ROW_LABEL = "R_Date"

# R_Dashboard row label -> this repo's facility category (config.FACILITY_CATEGORIES).
NETWORK_ROWS = {
    "R_Bike_Net_PBL": "protected",
    "R_Bike_Net_BBL": "buffered",
    "R_Bike_Net_BL": "painted",
    "R_Bike_Net_NG": "greenway",
    "R_Bike_Net_SL": "sharrow",
    "R_Bike_Net_Trail": "trail",
    "R_Bike_Net_Path": "other",       # "Access Path" — CDOT splits it out from trails
}
INSTALL_ROWS = {
    "R_Bike_Install_PBL": "protected",
    "R_Bike_Install_BBL": "buffered",
    "R_Bike_Install_BL": "painted",
    "R_Bike_Install_NG": "greenway",
    "R_Bike_Install_SL": "sharrow",
    "R_Bike_Install_Trail": "trail",
    "R_Bike_Install_Path": "other",
}
# CDOT's own totals, carried through so our sums can be checked against theirs
# rather than silently replacing them.
TOTAL_ROWS = {
    "network_on_street": "R_Bike_Net_On_T",
    "network_off_street": "R_Bike_Net_Off_T",
    "network_total": "R_Bike_Net_T",
    "network_growth": "R_Bike_Net_Growth",
    "installed_on_street": "R_Bike_Install_On_T",
    "installed_off_street": "R_Bike_Install_Off_T",
    "installed_total": "R_Bike_Install_T",
}
# Concrete upgrades of *existing* protected lanes. CDOT counts these inside its
# "miles installed" totals even though they add no new network mileage — which is
# why installed_total exceeds network_growth in upgrade-heavy years (2023: 50.42
# installed against 25.09 of growth). Published separately so the distinction is
# visible instead of buried.
CONCRETE_UPGRADE_ROW = "R_Bike_Install_PBLC"

# BIKE_DSPLY codes in the 2023+ layer schema are abbreviated, not the full labels
# used by the public layer's `displayrou` (so config.FACILITY_CATEGORY_MAP, which
# keys on those full labels, does not apply here).
SEGMENT_FACILITY_MAP = {
    "PROTECTED": "protected",
    "BUFFERED": "buffered",
    "BIKE": "painted",
    "NEIGHBORHOOD": "greenway",
    "SHARED": "sharrow",
}

# BW_INST_MO uses 0 and sentinel values (999, 9999) for "month not recorded".
UNKNOWN_MONTHS = {0, 999, 9999}


def _read_r_dashboard(path=None):
    """{row_label: {year: float}} from the dashboard's machine-readable sheet.

    First occurrence of a label wins — several `R_*` labels repeat further down
    the sheet to feed additional charts off the same numbers.
    """
    from openpyxl import load_workbook

    path = path or CDOT_COMPLETE_STREETS_DASHBOARD
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[DASHBOARD_SHEET]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    years = None
    for row in rows:
        if row and row[0] == YEAR_ROW_LABEL:
            years = [int(v) for v in row[1:] if isinstance(v, (int, float))]
            break
    if not years:
        raise ValueError(f"no '{YEAR_ROW_LABEL}' row in {path}:{DASHBOARD_SHEET}")

    out = {}
    for row in rows:
        label = row[0] if row else None
        if not isinstance(label, str) or label in out:
            continue
        values = {}
        for year, cell in zip(years, row[1:]):
            if isinstance(cell, (int, float)):
                values[year] = float(cell)
        if values:
            out[label] = values
    return out, years


def _series(table, years, row_map):
    """Per-year {category: miles} for one block of dashboard rows."""
    series = []
    for year in years:
        by_category = {c: 0.0 for c in FACILITY_CATEGORIES}
        present = False
        for label, category in row_map.items():
            value = table.get(label, {}).get(year)
            if value is not None:
                by_category[category] = round(value, 2)
                present = True
        if present:
            series.append({"year": year, "by_category": by_category,
                           "total": round(sum(by_category.values()), 2)})
    return series


def build_annual_series(path=None):
    """CDOT's 2010–2025 annual bikeway mileage — standing network and miles installed."""
    table, years = _read_r_dashboard(path)

    network = _series(table, years, NETWORK_ROWS)
    installed = _series(table, years, INSTALL_ROWS)

    upgrades = table.get(CONCRETE_UPGRADE_ROW, {})
    for point in installed:
        point["protected_concrete_upgrade"] = round(upgrades.get(point["year"], 0.0), 2)

    reported = {}
    for name, label in TOTAL_ROWS.items():
        values = table.get(label, {})
        if values:
            reported[name] = {str(y): round(v, 2) for y, v in sorted(values.items())}

    return {"years": [years[0], years[-1]], "network": network,
            "installed": installed, "cdot_reported_totals": reported}


def segment_install_years(path=None):
    """Centerline miles of the *2024* network grouped by the year each segment went in.

    Caveat, and it is the important one: this is the network that still existed in
    2024, attributed to install years. It is not the network as it stood in any past
    year. Lanes removed before 2024 are absent, and a lane upgraded in place carries
    its upgrade year, not its original one — which is exactly why CDOT's own buffered
    mileage *falls* after 2022 (115.6 -> 106.5) as buffered lanes became protected.
    Use `build_annual_series` for history; use this for segment-level questions like
    "when was the surviving protected network built".
    """
    import geopandas as gpd

    path = path or CDOT_BIKEWAY_2024_LAYER
    gdf = gpd.read_file(path)

    by_year = defaultdict(lambda: defaultdict(float))
    unknown_month_miles = 0.0
    for _, row in gdf.iterrows():
        year = row.get("BW_INST_YR")
        if not year or int(year) <= 0:
            continue
        category = SEGMENT_FACILITY_MAP.get(str(row.get("BIKE_DSPLY") or "").strip().upper(), "other")
        miles = float(row.get("ST_CL_MI") or 0.0)
        by_year[int(year)][category] += miles
        if int(row.get("BW_INST_MO") or 0) in UNKNOWN_MONTHS:
            unknown_month_miles += miles

    series = []
    for year in sorted(by_year):
        cats = {c: round(by_year[year].get(c, 0.0), 2) for c in FACILITY_CATEGORIES}
        series.append({"year": year, "by_category": cats,
                       "total": round(sum(cats.values()), 2)})

    total_miles = round(sum(p["total"] for p in series), 2)
    return {
        "layer": path.name if hasattr(path, "name") else str(path),
        "segments": len(gdf),
        "centerline_miles": total_miles,
        "miles_without_install_month": round(unknown_month_miles, 2),
        "note": (
            "Centerline miles of the network as it stood in CDOT's 2024 internal layer, "
            "grouped by per-segment install year (BW_INST_YR). Survivors only: lanes "
            "removed before 2024 do not appear, and upgraded lanes carry the upgrade "
            "year. Not a substitute for the annual series."
        ),
        "by_install_year": series,
    }


def build_cdot_bikeway_history(dashboard_path=None, layer_path=None):
    """The full committed document: CDOT's annual series plus the segment-level view."""
    annual = build_annual_series(dashboard_path)
    segments = segment_install_years(layer_path)
    first, last = annual["years"]
    return {
        "data_tier": "real",
        "source": "CDOT FOIA S145367-071326, released 2026-07-24",
        "source_records": [
            "data/foia/S145367-071326/records/CompleteStreets_Dashboard.xlsx (sheet R_Dashboard)",
            "data/foia/S145367-071326/records/GIS/Bikeway_Network_2024_Final.shp",
        ],
        "note": (
            f"CDOT's own bikeway mileage by facility type, {first}-{last}, from the Complete "
            "Streets program dashboard released under FOIA. `network` is the standing network "
            "at each year end; `installed` is miles added during that year. Note that CDOT "
            "counts concrete upgrades of existing protected lanes inside its installed totals "
            "(published separately here as protected_concrete_upgrade), so installed miles "
            "exceed network growth in upgrade-heavy years. Miles are centerline, matching "
            "bikeway_mileage_series.json's methodology."
        ),
        "annual": annual,
        "segment_install_years": segments,
    }


def main():
    doc = build_cdot_bikeway_history()
    CDOT_BIKEWAY_HISTORY_PATH.write_text(json.dumps(doc, indent=2) + "\n", encoding="utf-8")
    annual = doc["annual"]
    first, last = annual["years"]
    print(f"wrote {CDOT_BIKEWAY_HISTORY_PATH.relative_to(CDOT_BIKEWAY_HISTORY_PATH.parents[1])}")
    print(f"  annual series {first}-{last}: {len(annual['network'])} years")
    print(f"  segment install years: {doc['segment_install_years']['centerline_miles']} centerline mi")


if __name__ == "__main__":
    main()
